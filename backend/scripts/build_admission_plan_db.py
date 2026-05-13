# -*- coding: utf-8 -*-
"""Step 2: 从2025官方招生计划Excel构建完整招生数据库

生成 guangdong_2025_admission_plan.json，保留原始所有字段，
以 院校代码_专业组代码_专业代码 为唯一键。
"""
import json
import os
from pathlib import Path
from collections import OrderedDict
from openpyxl import load_workbook

EXCEL_PATH = r"C:\Users\18826\Desktop\20250618-广东-2025年-招生计划.xlsx"
DATA_DIR = Path(__file__).parent.parent / "data"
OUTPUT_FILE = DATA_DIR / "guangdong_2025_admission_plan.json"


def clean(val):
    if val is None:
        return ""
    return str(val).replace("\n", "").replace("\r", "").strip()


def main():
    print(f"[Step 2] 读取 Excel: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    headers = [clean(h) for h in header_row]
    print(f"  表头: {headers}")

    records = []
    seen_keys = set()
    duplicates = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        d = {headers[i]: clean(row[i]) if i < len(headers) and row[i] is not None else ""
             for i in range(len(headers))}

        uni_code = d.get("院校代码", "")
        group_code = d.get("专业组代码", "")
        major_code = d.get("专业代码", "")

        if not uni_code or not group_code:
            continue

        # 唯一键
        key = f"{uni_code}_{group_code}_{major_code}"
        if key in seen_keys:
            duplicates += 1
            continue
        seen_keys.add(key)

        # 数值字段转换
        try:
            plan_count = int(d.get("计划人数", 0)) if d.get("计划人数") else 0
        except ValueError:
            plan_count = 0
        try:
            duration = int(d.get("学制", 0)) if d.get("学制") else 0
        except ValueError:
            duration = 0
        try:
            tuition = float(d.get("学费", 0)) if d.get("学费") else 0
        except ValueError:
            tuition = 0

        records.append({
            "id": key,
            "year": int(d.get("年份", 2025)) if d.get("年份", "").isdigit() else 2025,
            "province": d.get("生源地", "广东"),
            "batch": d.get("批次", ""),
            "batch_note": d.get("批次备注", ""),
            "category": d.get("科类", ""),
            "university_code": uni_code,
            "university_name": d.get("院校名称", ""),
            "group_code": group_code,
            "major_code": major_code,
            "major_name": d.get("专业名称", ""),
            "major_note": d.get("专业备注", ""),
            "other_requirements": d.get("其他要求", ""),
            "subjects_requirement": d.get("选科要求", ""),
            "plan_count": plan_count,
            "duration": duration,
            "tuition": tuition,
        })

    wb.close()

    print(f"  总记录数: {len(records)}")
    print(f"  去重跳过: {duplicates}")

    os.makedirs(DATA_DIR, exist_ok=True)

    # 按年份和院校整理
    output = {
        "meta": {
            "source": "广东省教育考试院官方招生计划",
            "year": 2025,
            "generated": __import__("datetime").datetime.now().isoformat(),
        },
        "by_year": {
            "2025": records
        },
        "index": {
            "total": len(records),
            "university_count": len(set(r["university_code"] for r in records)),
            "group_count": len(set(f"{r['university_name']}_{r['group_code']}" for r in records)),
        }
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"  输出: {OUTPUT_FILE}")
    print("[Step 2] Done.")


if __name__ == "__main__":
    main()
