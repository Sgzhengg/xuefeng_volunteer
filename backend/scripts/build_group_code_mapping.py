# -*- coding: utf-8 -*-
"""Step 1: 从2025官方招生计划Excel构建专业组映射表

读取广东省教育考试院官方发布的《广东2025招生计划》，
按 院校名称_专业组代码 聚合，生成 group_code_to_majors.json
"""
import json
import os
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

EXCEL_PATH = r"C:\Users\18826\Desktop\20250618-广东-2025年-招生计划.xlsx"
DATA_DIR = Path(__file__).parent.parent / "data"
OUTPUT_FILE = DATA_DIR / "group_code_to_majors.json"


def clean(val):
    """清洗值：去除换行、前后空格"""
    if val is None:
        return ""
    return str(val).replace("\n", "").replace("\r", "").strip()


def main():
    print(f"[Step 1] 读取 Excel: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # 读取表头 (Row 2)
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    headers = [clean(h) for h in header_row]
    print(f"  表头: {headers}")

    # 按 (院校名称, 专业组代码) 聚合
    group_map = defaultdict(lambda: {
        "group_code": "",
        "category": "",
        "majors": [],
        "subjects": "",
        "plan_count": 0,
        "tuition": None,
        "duration": None,
        "university_code": "",
        "batch": "",
    })

    total_rows = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        d = {headers[i]: clean(row[i]) if i < len(headers) and row[i] is not None else ""
             for i in range(len(headers))}

        uni_name = d.get("院校名称", "")
        group_code = d.get("专业组代码", "")
        if not uni_name or not group_code:
            continue

        key = f"{uni_name}_{group_code}"
        entry = group_map[key]

        # 设置组级别字段（第一次遇到时）
        if not entry["group_code"]:
            entry["group_code"] = group_code
            entry["category"] = d.get("科类", "")
            entry["subjects"] = d.get("选科要求", "")
            entry["university_code"] = d.get("院校代码", "")
            entry["batch"] = d.get("批次", "")
            try:
                entry["duration"] = int(d.get("学制", 0)) if d.get("学制") else None
            except ValueError:
                entry["duration"] = None
            try:
                entry["tuition"] = float(d.get("学费", 0)) if d.get("学费") else None
            except ValueError:
                entry["tuition"] = None

        # 累加计划人数
        try:
            entry["plan_count"] += int(d.get("计划人数", 0)) if d.get("计划人数") else 0
        except ValueError:
            pass

        # 添加专业（去重）
        major_name = d.get("专业名称", "")
        major_code = d.get("专业代码", "")
        major_note = d.get("专业备注", "")
        if major_name:
            existing = [m for m in entry["majors"] if m["major_name"] == major_name]
            if not existing:
                entry["majors"].append({
                    "major_name": major_name,
                    "major_code": major_code,
                    "major_note": major_note,
                })
        total_rows += 1

    wb.close()
    print(f"  处理行数: {total_rows}")
    print(f"  生成映射数: {len(group_map)} 组")

    # 转换为 dict (用 key 显示)
    result = dict(group_map)

    # 统计
    uni_set = set()
    for k in result:
        parts = k.split("_", 1)
        if len(parts) > 1:
            uni_set.add(parts[0])

    print(f"  覆盖院校: {len(uni_set)} 所")
    print(f"  科类分布: 物理={sum(1 for v in result.values() if v['category']=='物理')}, "
          f"历史={sum(1 for v in result.values() if v['category']=='历史')}")

    # 保存
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"  输出: {OUTPUT_FILE} ({len(result)} entries)")
    print("[Step 1] Done.")


if __name__ == "__main__":
    main()
