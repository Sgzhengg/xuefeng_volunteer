"""
官方数据导入工具
支持从多种格式导入教育考试院官方数据：
- CSV文件
- Excel文件
- JSON文件
- PDF文件（需要额外库）
"""

import json
import csv
import re
from pathlib import Path
from typing import Dict, List, Optional, Union
from datetime import datetime


class OfficialDataImporter:
    """官方数据导入器"""

    def __init__(self):
        self.data_dir = Path("data")
        self.official_data_dir = self.data_dir / "official_data"
        self.official_data_dir.mkdir(exist_ok=True)

        # 加载院校数据
        with open(self.data_dir / "universities_list.json", 'r', encoding='utf-8') as f:
            uni_data = json.load(f)
        self.universities = {u["name"]: u for u in uni_data["universities"]}

        # 加载专业列表
        with open(self.data_dir / "majors_list.json", 'r', encoding='utf-8') as f:
            majors_data = json.load(f)
        self.majors = {m["name"]: m for m in majors_data["majors"]}

    def import_from_csv(self, csv_file: Union[str, Path], province: str, year: int) -> List[Dict]:
        """
        从CSV文件导入数据

        CSV格式要求：
        - 必须包含列：院校名称, 专业名称, 最低分, 最低位次
        - 可选列：平均分, 平均位次
        """
        csv_path = Path(csv_file)
        print(f"Importing data from {csv_path}...")

        data = []
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)

            for row_num, row in enumerate(reader, 1):
                try:
                    record = self._parse_row(row, province, year)
                    if record:
                        data.append(record)
                except Exception as e:
                    print(f"Warning: Failed to parse row {row_num}: {e}")

        print(f"Imported {len(data)} records from CSV")
        return data

    def import_from_excel(self, excel_file: Union[str, Path], province: str, year: int,
                         sheet_name: str = None) -> List[Dict]:
        """
        从Excel文件导入数据

        Excel格式要求：
        - 第一行：列名
        - 必须包含列：院校名称, 专业名称, 最低分, 最低位次
        """
        try:
            import openpyxl
        except ImportError:
            print("Error: openpyxl library not installed. Please install: pip install openpyxl")
            return []

        excel_path = Path(excel_file)
        print(f"Importing data from {excel_path}...")

        data = []
        wb = openpyxl.load_workbook(excel_path)

        # 如果指定了工作表，使用指定的工作表，否则使用第一个
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # 读取表头
        headers = [cell.value for cell in ws[1]]

        # 读取数据行
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                row_dict = dict(zip(headers, row))
                record = self._parse_row(row_dict, province, year)
                if record:
                    data.append(record)
            except Exception as e:
                print(f"Warning: Failed to parse row {row_num}: {e}")

        print(f"Imported {len(data)} records from Excel")
        return data

    def import_from_json(self, json_file: Union[str, Path], province: str, year: int) -> List[Dict]:
        """
        从JSON文件导入数据

        JSON格式：
        [
          {
            "university_name": "北京大学",
            "major_name": "计算机科学与技术",
            "min_score": 676,
            "min_rank": 1500
          }
        ]
        """
        json_path = Path(json_file)
        print(f"Importing data from {json_path}...")

        with open(json_path, 'r', encoding='utf-8') as f:
            raw_data = json.load(f)

        data = []
        for row_num, row in enumerate(raw_data, 1):
            try:
                record = self._parse_row(row, province, year)
                if record:
                    data.append(record)
            except Exception as e:
                print(f"Warning: Failed to parse record {row_num}: {e}")

        print(f"Imported {len(data)} records from JSON")
        return data

    def _parse_row(self, row: Dict, province: str, year: int) -> Optional[Dict]:
        """解析单行数据"""
        # 提取院校名称（支持多种列名）
        university_name = self._extract_value(row, ['院校名称', '院校', '学校', 'university', 'university_name'])
        if not university_name:
            return None

        # 提取专业名称（支持多种列名）
        major_name = self._extract_value(row, ['专业名称', '专业', 'major', 'major_name'])
        if not major_name:
            return None

        # 提取分数（支持多种列名）
        min_score = self._extract_number(row, ['最低分', '最低分数', 'min_score', 'score'])
        avg_score = self._extract_number(row, ['平均分', '平均分数', 'avg_score'])

        # 提取位次（支持多种列名）
        min_rank = self._extract_number(row, ['最低位次', '最低位', 'min_rank', 'rank'])
        avg_rank = self._extract_number(row, ['平均位次', '平均位', 'avg_rank'])

        # 数据验证
        if not min_score or not min_rank:
            return None

        # 查找院校信息
        university = self.universities.get(university_name)
        if not university:
            print(f"Warning: University not found in database: {university_name}")
            # 可以选择跳过或使用临时ID
            university_id = f"TEMP_{university_name}"
            university_level = ""
        else:
            university_id = university["id"]
            university_level = university.get("level", "")

        # 查找专业信息
        major = self.majors.get(major_name)
        if major:
            major_code = major["code"]
            major_category = major.get("category", "")
        else:
            # 如果找不到专业，尝试从专业名称推断
            major_code = self._infer_major_code(major_name)
            major_category = self._infer_major_category(major_name)

        # 生成university_major_id
        university_major_id = f"{university_id}_{major_code}"

        return {
            "university_major_id": university_major_id,
            "university_id": university_id,
            "university_name": university_name,
            "university_level": university_level,
            "major_code": major_code,
            "major_name": major_name,
            "major_category": major_category,
            "province": province,
            "year": year,
            "min_score": min_score,
            "avg_score": avg_score or min_score,
            "min_rank": min_rank,
            "avg_rank": avg_rank or min_rank,
            "data_source": f"official_{province}_{year}",
            "accuracy": "high"
        }

    def _extract_value(self, row: Dict, keys: List[str]) -> Optional[str]:
        """从字典中提取值（支持多个键名）"""
        for key in keys:
            value = row.get(key)
            if value:
                return str(value).strip()
        return None

    def _extract_number(self, row: Dict, keys: List[str]) -> Optional[int]:
        """从字典中提取数字"""
        for key in keys:
            value = row.get(key)
            if value:
                # 清理数字字符串
                if isinstance(value, str):
                    value = re.sub(r'[^\d]', '', value)
                try:
                    return int(float(value))
                except (ValueError, TypeError):
                    continue
        return None

    def _infer_major_code(self, major_name: str) -> str:
        """从专业名称推断专业代码"""
        # 简化实现：使用专业名称的哈希值
        return f"INFERRED_{hash(major_name) % 1000000:06d}"

    def _infer_major_category(self, major_name: str) -> str:
        """从专业名称推断专业类别"""
        # 简化实现：基于关键词匹配
        keywords = {
            "工学": ["计算机", "软件", "电子", "机械", "自动化", "电气"],
            "理学": ["数学", "物理", "化学", "生物"],
            "经济学": ["经济", "金融", "贸易"],
            "管理学": ["管理", "会计", "工商管理"],
            "医学": ["医学", "临床", "药学"],
            "法学": ["法学"],
            "文学": ["文学", "中文", "外语"],
            "历史学": ["历史"]
        }

        for category, kw_list in keywords.items():
            if any(kw in major_name for kw in kw_list):
                return category

        return "其他"

    def integrate_with_existing_data(self, new_data: List[Dict]) -> str:
        """将新数据集成到现有的major_rank_data.json"""
        print("\nIntegrating with existing major_rank_data.json...")

        # 加载现有数据
        with open(self.data_dir / "major_rank_data.json", 'r', encoding='utf-8') as f:
            existing_data = json.load(f)

        existing_records = existing_data["major_rank_data"]

        # 构建现有记录的唯一标识
        existing_keys = {
            (r["university_major_id"], r["province"], r["year"])
            for r in existing_records
        }

        # 添加新记录（官方数据优先，覆盖估算数据）
        added_count = 0
        updated_count = 0

        for record in new_data:
            key = (record["university_major_id"], record["province"], record["year"])

            if key in existing_keys:
                # 查找现有记录
                for i, existing_record in enumerate(existing_records):
                    if (existing_record["university_major_id"] == record["university_major_id"] and
                        existing_record["province"] == record["province"] and
                        existing_record["year"] == record["year"]):

                        # 如果现有数据是估算的，替换为官方数据
                        if existing_record.get("data_source", "").startswith("estimated"):
                            existing_records[i] = record
                            updated_count += 1
                        break
            else:
                # 新记录，直接添加
                existing_records.append(record)
                added_count += 1

        print(f"Updated {updated_count} existing records with official data")
        print(f"Added {added_count} new records")

        # 保存合并后的数据
        output_data = {
            "metadata": {
                "table_name": "major_rank_data",
                "description": "专业录取位次数据（官方+估算混合）",
                "version": "3.0.0",
                "generated_at": datetime.now().isoformat(),
                "total_records": len(existing_records),
                "accuracy": "混合（官方数据高精度，估算数据中等精度）"
            },
            "major_rank_data": existing_records
        }

        # 备份原文件
        backup_path = self.data_dir / f"major_rank_data_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(backup_path, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, ensure_ascii=False, indent=2)
        print(f"Backup saved to: {backup_path}")

        # 保存合并后的数据
        output_path = self.data_dir / "major_rank_data.json"
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)

        print(f"Merged data saved to: {output_path}")
        print(f"Total records after merge: {len(existing_records):,}")

        return str(output_path)

    def generate_import_report(self, imported_data: List[Dict], province: str, year: int):
        """生成导入报告"""
        print("\n" + "=" * 60)
        print("Data Import Report")
        print("=" * 60)

        print(f"\nSource: {province} ({year})")
        print(f"Total records imported: {len(imported_data)}")

        # 统计院校层次
        tier1_count = len([r for r in imported_data if r.get("university_level") == "985"])
        tier2_count = len([r for r in imported_data if r.get("university_level") == "211"])

        print(f"\nUniversity breakdown:")
        print(f"- 985 universities: {tier1_count} records")
        print(f"- 211 universities: {tier2_count} records")

        # 统计分数范围
        scores = [r["min_score"] for r in imported_data if r.get("min_score")]
        if scores:
            print(f"\nScore range:")
            print(f"- Min: {min(scores)}")
            print(f"- Max: {max(scores)}")
            print(f"- Average: {sum(scores) / len(scores):.1f}")

        # 统计位次范围
        ranks = [r["min_rank"] for r in imported_data if r.get("min_rank")]
        if ranks:
            print(f"\nRank range:")
            print(f"- Min: {min(ranks):,}")
            print(f"- Max: {max(ranks):,}")
            print(f"- Average: {sum(ranks) / len(ranks):,.0f}")

        print("\n" + "=" * 60)


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description='官方数据导入工具')
    parser.add_argument('--file', type=str, required=True,
                        help='数据文件路径（CSV/Excel/JSON）')
    parser.add_argument('--province', type=str, required=True,
                        help='省份名称（如：江苏）')
    parser.add_argument('--year', type=int, default=2024,
                        help='年份')
    parser.add_argument('--sheet', type=str,
                        help='Excel工作表名称（仅Excel文件）')

    args = parser.parse_args()

    print("=" * 60)
    print("Official Data Importer")
    print("=" * 60)

    try:
        importer = OfficialDataImporter()

        # 根据文件扩展名选择导入方法
        file_path = Path(args.file)
        file_ext = file_path.suffix.lower()

        if file_ext == '.csv':
            data = importer.import_from_csv(args.file, args.province, args.year)
        elif file_ext in ['.xlsx', '.xls']:
            data = importer.import_from_excel(args.file, args.province, args.year, args.sheet)
        elif file_ext == '.json':
            data = importer.import_from_json(args.file, args.province, args.year)
        else:
            print(f"Error: Unsupported file format: {file_ext}")
            return

        # 生成导入报告
        importer.generate_import_report(data, args.province, args.year)

        # 集成到现有数据
        if data:
            importer.integrate_with_existing_data(data)

            print("\n[SUCCESS] Data import completed successfully!")

    except Exception as e:
        print(f"\n[ERROR] Import failed: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
